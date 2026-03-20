# This script processes an Excel file containing trip data to identify and mark trips for deletion based on specific criteria:
# 1. The distance between the start and end points of the trip is less than or equal to 100 meters.
# 2. The duration of the trip is less than or equal to 5 minutes.
# The script adds a new column 'deleted' to indicate whether a trip should be deleted (1) or kept (0).
# Finally, it saves the modified data to a new Excel file while preserving the original formatting and highlighting the rows marked for deletion.
import pandas as pd
import numpy as np
from math import radians, cos, sin, asin, sqrt
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DISTANCE_THRESHOLD_METERS = 150
TIME_GAP_THRESHOLD_MINUTES = 8 



def haversine_distance(lat1, lon1, lat2, lon2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees).
    Returns distance in meters.
    """
    # Convert decimal degrees to radians 
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
    
    # Haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 
    r = 6371000  # Radius of earth in meters
    return c * r

def time_string_to_minutes(time_str):
    """
    Convert time string (HH:MM:SS or HH:MM:SS.ffffff) to minutes.
    Returns float value in minutes.
    """
    if pd.isna(time_str) or time_str is None:
        return 0
    
    time_str = str(time_str).strip()
    try:
        # Parse HH:MM:SS or HH:MM:SS.ffffff format
        parts = time_str.split(':')
        if len(parts) == 3:
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = float(parts[2])
            total_minutes = hours * 60 + minutes + seconds / 60
            return total_minutes
    except:
        pass
    return 0

def mark_deleted_trips(df):
    """
    Mark trips for deletion if:
    - Distance between start and end points <= DISTANCE_THRESHOLD_METERS
    - time_duration <= TIME_GAP_THRESHOLD_MINUTES
    
    Adds a new column 'deleted' with values:
    - 1: marked for deletion
    - 0: keep
    """
    df_copy = df.copy()
    
    # Initialize deleted column
    df_copy['deleted'] = 0
    
    for i in range(len(df_copy)):
        row = df_copy.iloc[i]
        
        try:
            # Get coordinates
            start_lat = float(row['start_latitude'])
            start_lon = float(row['start_longitude'])
            end_lat = float(row['end_latitude'])
            end_lon = float(row['end_longitude'])
            
            # Calculate distance between start and end
            distance = haversine_distance(start_lat, start_lon, end_lat, end_lon)
            
            # Get time duration in minutes
            if 'time_duration' in row:
                duration_minutes = time_string_to_minutes(row['time_duration'])
            else:
                duration_minutes = 0
            
            # Check if both conditions are met
            if distance <= DISTANCE_THRESHOLD_METERS and duration_minutes <= TIME_GAP_THRESHOLD_MINUTES:
                df_copy.loc[i, 'deleted'] = 1
                
        except Exception as e:
            # If there's any error (missing values, invalid data), skip this row
            continue
    
    return df_copy

def main():
    # File paths
    input_file = 'output/app_data_20250210_fill_purpose_adjust_mode_merged.xlsx'
    output_file = 'output/app_data_20250210_all_deleted_with_filled_purpose_with_adjusted_mode.xlsx'
    
    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Read the Excel file
    print(f"Reading file: {input_file}")
    df = pd.read_excel(input_file)
    
    print(f"Total rows: {len(df)}")
    print(f"Columns: {list(df.columns)}")
    
    # Mark trips for deletion
    print("\nMarking trips for deletion...")
    df_marked = mark_deleted_trips(df)
    
    deleted_count = df_marked['deleted'].sum()
    print(f"Found {deleted_count} trips marked for deletion")
    print(f"Criteria: distance <= {DISTANCE_THRESHOLD_METERS}m AND time_duration <= {TIME_GAP_THRESHOLD_MINUTES} minutes")
    
    # Save to Excel first (to get the data written)
    print(f"\nSaving to: {output_file}")
    df_marked.to_excel(output_file, index=False, engine='openpyxl')
    
    # Now preserve formatting from the original file
    print("Preserving original formatting and adding deleted row highlighting...")
    
    # Load both workbooks
    wb_original = load_workbook(input_file)
    ws_original = wb_original.active
    
    wb_output = load_workbook(output_file)
    ws_output = wb_output.active
    
    # Get the column index for 'deleted' column (new column)
    deleted_col_idx = None
    for idx, cell in enumerate(ws_output[1], 1):
        if cell.value == 'deleted':
            deleted_col_idx = idx
            break
    
    # Copy formatting from original file to output file
    # Start from row 2 (skip header)
    for row_idx in range(2, min(ws_original.max_row + 1, ws_output.max_row + 1)):
        # Copy formatting for all existing columns
        for col_idx in range(1, ws_original.max_column + 1):
            original_cell = ws_original.cell(row_idx, col_idx)
            output_cell = ws_output.cell(row_idx, col_idx)
            
            # Copy cell formatting
            if original_cell.has_style:
                output_cell.font = original_cell.font.copy()
                output_cell.fill = original_cell.fill.copy()
                output_cell.border = original_cell.border.copy()
                output_cell.alignment = original_cell.alignment.copy()
                output_cell.number_format = original_cell.number_format
    
    # Add highlighting for deleted rows (light red background)
    delete_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # Apply red highlighting to rows marked for deletion
    for row_idx in range(2, ws_output.max_row + 1):
        # Check if 'deleted' column is 1
        if deleted_col_idx and ws_output.cell(row_idx, deleted_col_idx).value == 1:
            # Highlight all cells in this row with light red
            for col_idx in range(1, ws_output.max_column + 1):
                ws_output.cell(row_idx, col_idx).fill = delete_fill
    
    wb_output.save(output_file)
    
    print(f"\nDone! Output file: {output_file}")
    print(f"Total rows: {len(df_marked)}")
    print(f"Rows marked for deletion: {deleted_count}")
    print(f"Rows to keep: {len(df_marked) - deleted_count}")
    print(f"Original formatting preserved and deleted rows highlighted in light red")

if __name__ == "__main__":
    main()