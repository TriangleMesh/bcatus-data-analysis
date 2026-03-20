import pandas as pd
import numpy as np
import json
from datetime import datetime, timedelta
from math import radians, cos, sin, asin, sqrt
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

DISTANCE_THRESHOLD_METERS = 700
TIME_GAP_THRESHOLD_MINUTES = 35    

def haversine_distance(lat1, lon1, lat2, lon2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees).
    Returns distance in meters.
    """
    # Convert decimal degrees to radians 
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
    
    # Haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 
    r = 6371000  # Radius of earth in meters
    return c * r

def time_string_to_seconds(time_str):
    """
    Convert time string (HH:MM:SS or HH:MM:SS.ffffff) to seconds.
    Returns float value in seconds.
    """
    if pd.isna(time_str) or time_str is None:
        return 0
    
    time_str = str(time_str).strip()
    try:
        # Parse HH:MM:SS or HH:MM:SS.ffffff format
        parts = time_str.split(':')
        if len(parts) == 3:
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = float(parts[2])
            total_seconds = hours * 3600 + minutes * 60 + seconds
            return total_seconds
    except:
        pass
    return 0

def seconds_to_time_string(seconds):
    """
    Convert seconds (float) back to HH:MM:SS.ffffff format.
    """
    if seconds == 0:
        return "0:00:00"
    
    total_seconds = int(seconds)
    microseconds = int((seconds - total_seconds) * 1000000)
    
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    secs = total_seconds % 60
    
    if microseconds > 0:
        return f"{hours}:{minutes:02d}:{secs:02d}.{microseconds:06d}"
    else:
        return f"{hours}:{minutes:02d}:{secs:02d}"

def parse_routes(routes_data):
    """
    Parse routes data - could be JSON string with single quotes (needs conversion),
    list, or other format. Returns the parsed data or None.
    """
    if pd.isna(routes_data) or routes_data is None:
        return None
    
    if isinstance(routes_data, list):
        return routes_data
    
    if isinstance(routes_data, str):
        try:
            # Try standard JSON parsing first
            return json.loads(routes_data)
        except:
            try:
                # If that fails, try converting single quotes to double quotes
                # This is a common format for Python data structures
                routes_data_fixed = routes_data.replace("'", '"')
                return json.loads(routes_data_fixed)
            except:
                return None
    
    return None

def mark_merge_candidates(df):
    """
    Mark potential merge candidates in a separate column before merging.
    This helps identify which trips could be merged based on:
    - Same accessCode, start_date, end_date
    - Time gap <= TIME_GAP_THRESHOLD_MINUTES
    - Geographic distance <= DISTANCE_THRESHOLD_METERS
    - Same purpose_of_travel (both must have same non-blank purpose)
    
    Returns dataframe with new column 'merge_candidate' with three values:
    - 0: Not part of any merge (default)
    - 1: Small trip that will be merged into a longer trip
    - 2: New long trip created from merging multiple small trips
    """
    df_copy = df.copy()
    
    # Convert time columns to datetime for comparison
    df_copy['end_time_dt'] = pd.to_datetime(df_copy['end_time'], errors='coerce')
    df_copy['start_time_dt'] = pd.to_datetime(df_copy['start_time'], errors='coerce')
    
    # Initialize merge_candidate column
    df_copy['merge_candidate'] = 0
    
    # First pass: identify all consecutive merge chains
    i = 0
    while i < len(df_copy):
        merge_chain = [i]  # Start with current index
        
        # Look ahead to find all consecutive mergeable trips
        j = i
        while j < len(df_copy) - 1:
            current = df_copy.iloc[j]
            next_row = df_copy.iloc[j + 1]
            
            # Check grouping: same accessCode, start_date, end_date
            if (current['accessCode'] != next_row['accessCode'] or
                current['start_date'] != next_row['start_date'] or
                current['end_date'] != next_row['end_date']):
                break
            
            # Check time gap: <= TIME_GAP_THRESHOLD_MINUTES
            if pd.isna(current['end_time_dt']) or pd.isna(next_row['start_time_dt']):
                break
            
            time_diff = (next_row['start_time_dt'] - current['end_time_dt']).total_seconds() / 60
            if time_diff < 0 or time_diff > TIME_GAP_THRESHOLD_MINUTES:
                break
            
            # Check location distance: <= DISTANCE_THRESHOLD_METERS
            try:
                current_end_lat = float(current['end_latitude'])
                current_end_lon = float(current['end_longitude'])
                next_start_lat = float(next_row['start_latitude'])
                next_start_lon = float(next_row['start_longitude'])
                
                distance = haversine_distance(current_end_lat, current_end_lon,
                                             next_start_lat, next_start_lon)
                if distance > DISTANCE_THRESHOLD_METERS:
                    break
            except:
                break
            
            # Check same purpose: trips must have the same purpose to be merged
            # if 'purpose_of_travel' in df_copy.columns:
            #     current_purpose = current['purpose_of_travel']
            #     next_purpose = next_row['purpose_of_travel']
            #     # Both must have the same purpose (not blank/NaN)
            #     if pd.isna(current_purpose) or pd.isna(next_purpose):
            #         break
            #     if current_purpose != next_purpose:
            #         break
            
            # This trip can be merged with the next one
            merge_chain.append(j + 1)
            j += 1
        
        # If we found a merge chain (more than 1 trip), mark all small trips as 1
        if len(merge_chain) > 1:
            for idx in merge_chain:
                df_copy.loc[idx, 'merge_candidate'] = 1  # Mark small trips as 1
        
        # Move to the next unprocessed trip
        i = j + 1
    
    # Drop temporary datetime columns
    df_copy = df_copy.drop(['end_time_dt', 'start_time_dt'], axis=1)
    
    return df_copy

def merge_trips(df):
    """
    Merge consecutive trips based on:
    - Same accessCode, start_date, end_date
    - Time gap between end_time and next start_time < 5 minutes
    - Geographic distance < 100 meters
    
    Returns original dataframe with merged rows inserted after source rows.
    Keeps all original rows and adds merged rows after them.
    """
    
    # Convert time columns to datetime for faster comparison
    df_copy = df.copy()
    df_copy['end_time'] = pd.to_datetime(df_copy['end_time'], errors='coerce')
    df_copy['start_time'] = pd.to_datetime(df_copy['start_time'], errors='coerce')
    
    result_rows = []
    processed = set()
    
    for i in range(len(df_copy)):
        if i in processed:
            continue
        
        # Add original row
        current_row = df.iloc[i].copy()
        # Initialize merge columns if they don't exist
        if 'is_merged' not in current_row:
            current_row['is_merged'] = False
        if 'merged_source_rows' not in current_row:
            current_row['merged_source_rows'] = ''
        result_rows.append((current_row, False, None))
        
        # Find consecutive rows to merge
        merged_indices = [i]
        last_idx = i
        
        # Look ahead for mergeable rows
        for j in range(i + 1, len(df_copy)):
            if j in processed:
                break  # Stop if we hit a processed row
            
            current = df_copy.iloc[i]
            next_row = df_copy.iloc[j]
            prev_row = df_copy.iloc[last_idx]
            
            # Check grouping: same accessCode, start_date, end_date
            if (current['accessCode'] != next_row['accessCode'] or
                current['start_date'] != next_row['start_date'] or
                current['end_date'] != next_row['end_date']):
                break
            
            # Check time gap: <= 5 minutes
            if pd.isna(prev_row['end_time']) or pd.isna(next_row['start_time']):
                break
            
            time_diff = (next_row['start_time'] - prev_row['end_time']).total_seconds() / 60
            if time_diff < 0 or time_diff > TIME_GAP_THRESHOLD_MINUTES:
                break
            
            # Check location distance: <= 100 meters
            try:
                prev_end_lat = float(prev_row['end_latitude'])
                prev_end_lon = float(prev_row['end_longitude'])
                next_start_lat = float(next_row['start_latitude'])
                next_start_lon = float(next_row['start_longitude'])
                
                distance = haversine_distance(prev_end_lat, prev_end_lon,
                                             next_start_lat, next_start_lon)
                if distance > DISTANCE_THRESHOLD_METERS:
                    break
            except:
                break
            
            # Check same purpose: trips must have the same purpose to be merged
            # If purpose_of_travel column exists, check that both trips have the same purpose
            # if 'purpose_of_travel' in df_copy.columns:
            #     prev_purpose = prev_row['purpose_of_travel']
            #     next_purpose = next_row['purpose_of_travel']
            #     # Both must have the same purpose (not blank/NaN)
            #     if pd.isna(prev_purpose) or pd.isna(next_purpose):
            #         break
            #     if prev_purpose != next_purpose:
            #         break
            
            # Can merge this row
            merged_indices.append(j)
            last_idx = j
        
        # If found rows to merge
        if len(merged_indices) > 1:
            # Add all original rows to be merged
            for idx in merged_indices[1:]:
                original_row = df.iloc[idx].copy()
                if 'is_merged' not in original_row:
                    original_row['is_merged'] = False
                if 'merged_source_rows' not in original_row:
                    original_row['merged_source_rows'] = ''
                result_rows.append((original_row, False, None))
                processed.add(idx)
            
            # Create and add merged row
            merged_row = create_merged_row(df, merged_indices, current_row)
            merge_info = f"Merged from rows: {', '.join([str(idx+2) for idx in merged_indices])}"
            result_rows.append((merged_row, True, merge_info))
        
        processed.add(i)
    
    return result_rows

def create_merged_row(df, indices, first_row):
    """
    Create a merged row from multiple consecutive rows.
    - Combines time_duration by summing all durations (converts HH:MM:SS to seconds, then back)
    - Combines routes into a 2D array structure (JSON format)
    - Takes first row's start info, last row's end info
    - Marks with merge source information
    """
    merged_row = first_row.copy()
    
    last_row = df.iloc[indices[-1]]
    
    # Update end information from the last row
    merged_row['end_time'] = last_row['end_time']
    merged_row['end_latitude'] = last_row['end_latitude']
    merged_row['end_longitude'] = last_row['end_longitude']
    
    # Sum up all time_duration values by converting to seconds first
    total_seconds = 0
    for idx in indices:
        row = df.iloc[idx]
        if 'time_duration' in row:
            duration_str = row['time_duration']
            total_seconds += time_string_to_seconds(duration_str)
    
    # Convert back to time string format (HH:MM:SS or HH:MM:SS.ffffff)
    merged_row['time_duration'] = seconds_to_time_string(total_seconds)
    
    # Merge routes into a 2D array structure
    routes_array = []
    for idx in indices:
        row = df.iloc[idx]
        if 'routes' in row:
            routes_data = parse_routes(row['routes'])
            if routes_data is not None:
                routes_array.append(routes_data)
    
    # Convert routes array to JSON string for Excel storage
    if routes_array:
        merged_row['routes'] = json.dumps(routes_array, ensure_ascii=False)
    else:
        merged_row['routes'] = None
    
    # Mark as merged and record source row numbers (Excel row numbers, i.e., index + 2)
    merged_row['is_merged'] = True
    source_rows_str = ', '.join([str(idx + 2) for idx in indices])
    merged_row['merged_source_rows'] = f"Merged from rows: {source_rows_str}"
    
    # Mark this merged row as type 2 (new long trip created from merging)
    merged_row['merge_candidate'] = 2
    
    return merged_row

def main():
    # File paths
    input_file = 'output/app_data_20250210_fill_purpose_adjust_mode.xlsx'
    output_file = 'output/app_data_20250210_fill_purpose_adjust_mode_merged.xlsx'

    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Read the Excel file
    print(f"Reading file: {input_file}")
    df = pd.read_excel(input_file)
    
    print(f"Total rows: {len(df)}")
    print(f"Columns: {list(df.columns)}")
    
    # Check for required columns
    required_cols = ['accessCode', 'start_date', 'end_date', 'start_time', 'end_time',
                     'start_latitude', 'start_longitude', 'end_latitude', 'end_longitude']
    missing_cols = [col for col in required_cols if col not in df.columns]
    
    if missing_cols:
        print(f"WARNING: Missing columns: {missing_cols}")
        print("Available columns:", list(df.columns))
    
    # Mark potential merge candidates first
    print("\nMarking potential merge candidates...")
    df = mark_merge_candidates(df)
    
    # Count small trips that will be merged (merge_candidate = 1)
    small_trips_count = (df['merge_candidate'] == 1).sum()
    print(f"Found {small_trips_count} small trips that will be merged into longer trips")
    
    # Merge trips
    print("\nMerging trips...")
    import sys
    
    result_rows = merge_trips(df)
    
    print(f"Merge complete. Processing results...")
    
    # Create new dataframe with all rows (original + merged)
    new_rows = [row_data for row_data, is_merged, info in result_rows]
    new_df = pd.DataFrame(new_rows)
    
    # Count the three types of trips
    type_0_count = (new_df['merge_candidate'] == 0).sum()  # Not merged
    type_1_count = (new_df['merge_candidate'] == 1).sum()  # Small trips that were merged
    type_2_count = (new_df['merge_candidate'] == 2).sum()  # New long trips created
    
    print(f"\n" + "="*60)
    print("MERGE CANDIDATE STATISTICS:")
    print("="*60)
    print(f"Type 0 (Not merged): {type_0_count} trips")
    print(f"Type 1 (Small trips merged): {type_1_count} trips")
    print(f"Type 2 (New long trips created): {type_2_count} trips")
    print(f"Total trips in output: {len(new_df)}")
    print("="*60)
    
    # Save to Excel with formatting
    print(f"\nSaving to: {output_file}")
    new_df.to_excel(output_file, index=False, engine='openpyxl')
    
    # Format the Excel file with highlighting for different merge types
    print("Applying formatting to rows based on merge_candidate type...")
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Create highlight colors for different types
    # Type 1: Light blue for small trips that will be merged
    type1_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    # Type 2: Yellow for new long trips created from merging
    type2_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    type2_font = Font(bold=True, italic=True, color="FF0000")
    
    # Find the column index for 'merge_candidate'
    merge_candidate_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'merge_candidate':
            merge_candidate_col_idx = idx
            break
    
    # Apply formatting based on merge_candidate value
    if merge_candidate_col_idx:
        row_num = 2  # Start from row 2 (after header)
        for row_data, is_merged, info in result_rows:
            merge_type = row_data.get('merge_candidate', 0)
            
            if merge_type == 1:
                # Light blue for small trips (Type 1)
                for cell in ws[row_num]:
                    cell.fill = type1_fill
            elif merge_type == 2:
                # Yellow with red bold italic font for new long trips (Type 2)
                for cell in ws[row_num]:
                    cell.fill = type2_fill
                    cell.font = type2_font
            
            row_num += 1
    else:
        # Fallback to old method if merge_candidate column not found
        row_num = 2
        for row_data, is_merged, info in result_rows:
            if is_merged and info:
                for cell in ws[row_num]:
                    cell.fill = type2_fill
                    cell.font = type2_font
            row_num += 1
    
    wb.save(output_file)
    
    print(f"\nDone! Output file: {output_file}")
    print(f"Original rows: {len(df)}")
    print(f"New rows (with merged): {len(new_df)}")
    print(f"Merged rows added: {len(new_df) - len(df)}")

if __name__ == "__main__":
    main()
